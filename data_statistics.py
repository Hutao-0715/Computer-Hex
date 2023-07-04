from openpyxl import load_workbook
# pip install openpyxl
#def
def get_attr(line): # return a str arr
    start = 0
    end = 0
    res = []
    for i in range(len(line)):
        if line[i] == "=": start = i
        elif line[i]==" " and start!=0:
            end = i
            res.append(line[start+1:end])
    return res
def find_pos(set, sheet):
    c = 1
    r = 4
    while 1:
        if sheet.cell(row=2,column=c).value == None:
            sheet.cell(row=2,column=c,value=float(set[0]))
            sheet.cell(row=2,column=c+1,value=int(set[1]))
            sheet.cell(row=2,column=c+2,value=set[2])
            break
        if  (sheet.cell(row=2,column=c).value >= float(set[0])-0.01 and sheet.cell(row=2,column=c).value <= float(set[0])+0.01) and sheet.cell(row=2,column=c+1).value == int(set[1]) and sheet.cell(row=2,column=c+2).value == set[2]:
            break
        c += 3
    while 1:
        if sheet.cell(row=r,column=c).value == None:
            break
        r += 1
    return [r,c]

#start
statistics_path = f"D:\desktop\程式設計\專題\報告用資料\MCTS-SaveBridge-NoMinus結果統計.xlsx"
data_path = f"D:\desktop\data\MCTS-SaveBridge-NoMinus_C_1.1-1.9_s_60000_L_2.txt" #已經跑過了

wb = load_workbook(statistics_path)
sheet = wb[f"level2"]

data_start = 0
with open(data_path) as data:
    for line in data.readlines():
        if line.startswith("<experiment"):
            data_start = 1
        if line.startswith("<set>"): # and data_start:
            data_start = 0
            s = get_attr(line)
            pos = find_pos(s,sheet)
            cur_r = pos[0]
            cur_c = pos[1]
        if line.startswith("<resault>"):
            s = get_attr(line)
            sheet.cell(row=cur_r,column=cur_c,value=int(s[0]))
            sheet.cell(row=cur_r,column=cur_c+1,value=float(s[1]))
            sheet.cell(row=cur_r,column=cur_c+2,value=int(s[2]))
            cur_r += 1

wb.save(statistics_path)
